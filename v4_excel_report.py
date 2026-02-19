# v4_excel_report.py
# Excel workbook writer - multi-tab formatted output
# =============================================================================

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


# Styling constants
HEADER_FILL = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Calibri", size=10)
DATA_ALIGN = Alignment(vertical="center")
CURRENCY_FORMAT = '#,##0'
CURRENCY_CENTS_FORMAT = '#,##0.00'
PERCENT_FORMAT = '0.0"%"'
NUMBER_FORMAT = '#,##0'

# Columns that should show decimal cents (Avg, Median, per-unit metrics)
_CENTS_KEYWORDS = {
    "avg", "average", "median", "per employee", "per account",
    "per card", "ticket", "revenue at risk", "winback revenue",
    "est. annual revenue", "first 3m avg", "last 3m avg",
    "avg monthly", "interchange",
}
THIN_BORDER = Border(
    bottom=Side(style="thin", color="E2E8F0"),
)
ALT_ROW_FILL = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")


def generate_excel_report(storyline_results: dict, config: dict, output_path: str):
    """
    Generate a multi-tab Excel workbook.

    Parameters
    ----------
    storyline_results : dict
        Keys are storyline names. Values are dicts with:
            'title': str
            'sheets': list of dicts, each with:
                'name': str (sheet tab name, max 31 chars)
                'df': pd.DataFrame
                'currency_cols': list of str (columns to format as currency)
                'pct_cols': list of str (columns to format as percentage)
                'number_cols': list of str (columns to format with comma separators)
    config : dict
    output_path : str
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Order keys: s0_executive first, then the rest in original order
    ordered_keys = []
    if "s0_executive" in storyline_results:
        ordered_keys.append("s0_executive")
    ordered_keys.extend(k for k in storyline_results if k != "s0_executive")

    sheet_count = 0
    for key in ordered_keys:
        result = storyline_results[key]
        for sheet_info in result.get("sheets", []):
            sheet_name = sheet_info["name"][:31]
            df = sheet_info["df"]
            if df is None or df.empty:
                continue

            ws = wb.create_sheet(title=sheet_name)
            sheet_count += 1

            # Write title row
            ws.merge_cells(
                start_row=1,
                start_column=1,
                end_row=1,
                end_column=max(len(df.columns), 1),
            )
            title_cell = ws.cell(row=1, column=1, value=result["title"])
            title_cell.font = Font(name="Calibri", size=14, bold=True, color="2E4057")
            title_cell.alignment = Alignment(horizontal="left")

            # Write subtitle
            ws.merge_cells(
                start_row=2,
                start_column=1,
                end_row=2,
                end_column=max(len(df.columns), 1),
            )
            subtitle = ws.cell(
                row=2,
                column=1,
                value=f"{sheet_info.get('subtitle', sheet_name)} | {config.get('client_name', '')}",
            )
            subtitle.font = Font(name="Calibri", size=10, italic=True, color="8B95A2")

            # Blank row
            start_row = 4

            # Write headers
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=col_name)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = HEADER_ALIGN
                cell.border = Border(
                    bottom=Side(style="medium", color="2E4057")
                )

            # Write data rows
            currency_cols = set(sheet_info.get("currency_cols", []))
            pct_cols = set(sheet_info.get("pct_cols", []))
            number_cols = set(sheet_info.get("number_cols", []))

            for row_idx, (_, row) in enumerate(df.iterrows(), start_row + 1):
                for col_idx, (col_name, value) in enumerate(row.items(), 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = DATA_FONT
                    cell.alignment = DATA_ALIGN
                    cell.border = THIN_BORDER

                    # Apply number formats
                    if col_name in currency_cols:
                        col_lower = col_name.lower()
                        if any(kw in col_lower for kw in _CENTS_KEYWORDS):
                            cell.number_format = CURRENCY_CENTS_FORMAT
                        else:
                            cell.number_format = CURRENCY_FORMAT
                    elif col_name in pct_cols:
                        cell.number_format = PERCENT_FORMAT
                    elif col_name in number_cols:
                        cell.number_format = NUMBER_FORMAT

                    # Alternating row fill
                    if (row_idx - start_row) % 2 == 0:
                        cell.fill = ALT_ROW_FILL

            # Auto-fit column widths
            for col_idx in range(1, len(df.columns) + 1):
                col_letter = get_column_letter(col_idx)
                max_len = len(str(df.columns[col_idx - 1]))
                for row in ws.iter_rows(
                    min_row=start_row + 1,
                    max_row=ws.max_row,
                    min_col=col_idx,
                    max_col=col_idx,
                ):
                    for cell in row:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

            # Freeze panes (headers visible when scrolling)
            ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # Add overview sheet at the beginning
    _add_overview_sheet(wb, storyline_results, config, sheet_count)

    # Save
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"  Excel report: {output} ({sheet_count} sheets)")


def _add_overview_sheet(wb, storyline_results, config, sheet_count):
    """Add a summary overview as the first sheet."""
    ws = wb.create_sheet(title="Overview", index=0)

    # Title
    ws.merge_cells("A1:D1")
    title = ws.cell(row=1, column=1, value=f"{config.get('client_name', '')} - Transaction Analysis")
    title.font = Font(name="Calibri", size=18, bold=True, color="2E4057")

    ws.merge_cells("A2:D2")
    from datetime import datetime
    subtitle = ws.cell(
        row=2,
        column=1,
        value=f"Client ID: {config.get('client_id', '')} | Generated: {datetime.now().strftime('%B %d, %Y')}",
    )
    subtitle.font = Font(name="Calibri", size=11, italic=True, color="8B95A2")

    # Table of contents
    row = 4
    ws.cell(row=row, column=1, value="Table of Contents").font = Font(
        name="Calibri", size=14, bold=True, color="2E4057"
    )
    row += 1

    headers = ["Storyline", "Sheets", "Description"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    row += 1
    overview_keys = []
    if "s0_executive" in storyline_results:
        overview_keys.append("s0_executive")
    overview_keys.extend(k for k in storyline_results if k != "s0_executive")
    for key in overview_keys:
        result = storyline_results[key]
        sheet_names = [s["name"] for s in result.get("sheets", [])]
        ws.cell(row=row, column=1, value=result["title"]).font = Font(
            name="Calibri", size=10, bold=True
        )
        ws.cell(row=row, column=2, value=", ".join(sheet_names)).font = DATA_FONT
        ws.cell(row=row, column=3, value=result.get("description", "")).font = DATA_FONT
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 60


def format_df_for_excel(df: pd.DataFrame, currency_cols=None, pct_cols=None) -> pd.DataFrame:
    """
    Prepare a DataFrame for Excel export.
    Ensures numeric columns have proper types (not formatted strings).
    """
    result = df.copy()
    for col in result.columns:
        if result[col].dtype == "object":
            # Try to convert string currency/pct back to numeric
            sample = result[col].dropna().head(5)
            if len(sample) > 0:
                first = str(sample.iloc[0])
                if first.startswith("$"):
                    result[col] = (
                        result[col]
                        .astype(str)
                        .str.replace("$", "", regex=False)
                        .str.replace(",", "", regex=False)
                    )
                    result[col] = pd.to_numeric(result[col], errors="coerce")
                elif first.endswith("%"):
                    result[col] = (
                        result[col]
                        .astype(str)
                        .str.replace("%", "", regex=False)
                    )
                    result[col] = pd.to_numeric(result[col], errors="coerce")
    return result
